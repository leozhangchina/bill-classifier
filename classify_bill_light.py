#!/usr/bin/env python3
"""微信/支付宝账单轻量分类器：本地规则分类，无大模型、无交互确认。"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

if sys.platform == "win32":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请运行：python -m pip install openpyxl") from exc


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CATEGORIES = SCRIPT_DIR / "categories.default.json"
DEFAULT_OVERRIDES = SCRIPT_DIR / "merchant-overrides.json"
APPENDED_HEADERS = ["最终分类", "是否需要人工确认"]

# 微信和支付宝采用不同列名，这里统一为程序内部字段。
HEADER_ALIASES = {
    "time": ["交易时间", "时间", "支付时间", "记账时间"],
    "type": ["交易类型", "交易分类", "类型"],
    "counterparty": ["交易对方", "商户名称", "对方", "收款方"],
    "account": ["对方账号", "交易对方账号"],
    "product": ["商品", "商品说明", "商品名称", "摘要", "交易说明"],
    "direction": ["收/支", "收支", "交易方向"],
    "amount": ["金额(元)", "金额（元）", "金额", "交易金额"],
    "payment": ["支付方式", "付款方式", "收/付款方式", "收付款方式"],
    "status": ["当前状态", "交易状态", "状态"],
    "transaction_id": ["交易单号", "交易订单号", "交易号", "订单号"],
    "merchant_id": ["商户单号", "商家订单号", "商家订单号"],
    "remark": ["备注", "附言", "说明"],
}

# 对两个样本中常见、但通用分类表尚未覆盖的表达做少量补充。
# 用户仍可直接编辑 categories.default.json 的 keywords 动态扩充规则。
EXTRA_KEYWORDS = {
    "食品": ["餐饮美食", "一餐", "美食", "小杨生煎", "萨莉亚", "肯德基", "麦当劳", "汉堡王", "麻辣烫", "螺蛳粉"],
    "饮料": ["coffee", "manner"],
    "水果零食": ["果优鲜", "油桃", "蟠桃"],
    "家居用品": ["日用百货"],
    "电子产品": ["充电宝"],
    "私家车费用": ["爱车养车", "充电订单", "停车场"],
    "运动健身": ["体育系", "体育场", "场地费"],
    "会员订阅": ["连续包月", "自动续费"],
    "利息收入": ["收益发放", "余额宝收益"],
    "信用卡还款": ["花呗还款", "花呗主动还款"],
    "存款取款": ["转入零钱通", "转出零钱通", "转入余额宝", "余额宝转入", "转账收款到余额宝"],
}

# 支付宝“交易分类”可作为辅助证据，但不会覆盖更具体的商品关键词。
SOURCE_CATEGORY_HINTS = {
    "餐饮美食": "食品",
    "日用百货": "家居用品",
    "爱车养车": "私家车费用",
    "信用借还": "信用卡还款",
    "转账红包": "其他转账",
}


@dataclass
class Category:
    label: str
    flow: str
    keywords: list[str] = field(default_factory=list)


@dataclass
class Config:
    categories: list[Category]
    fallback_by_flow: dict[str, str]
    merchant_rules: list[dict[str, Any]]

    @property
    def labels(self) -> list[str]:
        return [item.label for item in self.categories]

    def categories_for_flow(self, flow: str) -> list[Category]:
        return [item for item in self.categories if item.flow == flow]

    def labels_for_flow(self, flow: str) -> set[str]:
        return {item.label for item in self.categories_for_flow(flow)}


@dataclass
class SourceData:
    headers: list[str]
    lookup: dict[str, int]
    rows: list[list[Any]]
    sheet_name: str
    provider: str
    removed_leading_rows: int


@dataclass
class Result:
    category: str
    confidence: float
    method: str


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.lstrip("`").replace("\u3000", " ")).strip().lower()


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s:：]", "", normalize_text(value))


def header_lookup(headers: list[Any]) -> dict[str, int]:
    normalized = [normalize_header(item) for item in headers]
    result: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        choices = {normalize_header(alias) for alias in aliases}
        result[key] = next((index for index, item in enumerate(normalized) if item in choices), -1)
    return result


def header_score(row: list[Any] | tuple[Any, ...]) -> int:
    return sum(index >= 0 for index in header_lookup(list(row)).values())


def find_header_row(rows: list[list[Any]]) -> int:
    score, index = max(
        ((header_score(row), index) for index, row in enumerate(rows[:100])),
        default=(-1, -1),
    )
    if score < 6:
        raise ValueError(f"无法识别微信/支付宝账单表头（最多匹配 {score} 个字段）")
    return index


def decode_csv(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别 CSV 编码")


def read_rows(path: Path, requested_sheet: str | None) -> tuple[list[list[Any]], str]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if requested_sheet:
                if requested_sheet not in workbook.sheetnames:
                    raise ValueError(f"不存在工作表：{requested_sheet}")
                sheet = workbook[requested_sheet]
            else:
                sheet = max(
                    workbook.worksheets,
                    key=lambda item: max(
                        (header_score(row) for row in item.iter_rows(min_row=1, max_row=min(100, item.max_row), values_only=True)),
                        default=-1,
                    ),
                )
            return [list(row) for row in sheet.iter_rows(values_only=True)], sheet.title
        finally:
            workbook.close()

    if path.suffix.lower() in {".csv", ".tsv"}:
        text = decode_csv(path.read_bytes())
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        if path.suffix.lower() == ".csv":
            try:
                delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|").delimiter
            except csv.Error:
                pass
        return [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)], "账单"

    raise ValueError("仅支持 .xlsx、.csv 或 .tsv")


def parse_datetime(value: Any) -> Any:
    if isinstance(value, datetime) or not isinstance(value, str):
        return value
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    return value.strip()


def parse_amount(value: Any) -> Any:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    match = re.search(r"-?\d+(?:\.\d+)?", normalize_text(value).replace(",", "").replace("¥", "").replace("￥", ""))
    return float(match.group(0)) if match else value


def clean_source(path: Path, requested_sheet: str | None) -> SourceData:
    raw_rows, sheet_name = read_rows(path, requested_sheet)
    header_index = find_header_row(raw_rows)
    header_values = list(raw_rows[header_index])
    while header_values and not normalize_text(header_values[-1]):
        header_values.pop()
    headers = [str(value).strip() if normalize_text(value) else f"未命名列{index + 1}" for index, value in enumerate(header_values)]
    lookup = header_lookup(headers)
    provider = "支付宝" if "支付宝" in path.name or "交易分类" in headers else "微信"
    rows: list[list[Any]] = []
    for raw in raw_rows[header_index + 1:]:
        row = list(raw[:len(headers)]) + [None] * max(0, len(headers) - len(raw))
        if all(not normalize_text(value) for value in row):
            continue
        identity = [lookup[key] for key in ("time", "counterparty", "product", "amount") if lookup[key] >= 0]
        if identity and all(not normalize_text(row[index]) for index in identity):
            continue
        if lookup["time"] >= 0:
            row[lookup["time"]] = parse_datetime(row[lookup["time"]])
        if lookup["amount"] >= 0:
            row[lookup["amount"]] = parse_amount(row[lookup["amount"]])
        # 支付宝原始账单把余额宝每日收益标成“不计收支”，但它实际增加资产，
        # 因此在输出清洗后的表格时统一修正为“收入”。
        product_text = normalize_text(row[lookup["product"]]) if lookup["product"] >= 0 else ""
        if (
            provider == "支付宝"
            and lookup["direction"] >= 0
            and "余额宝" in product_text
            and "收益发放" in product_text
        ):
            row[lookup["direction"]] = "收入"
        for key in ("transaction_id", "merchant_id"):
            index = lookup[key]
            if index >= 0 and row[index] is not None:
                row[index] = str(row[index]).strip().lstrip("`")
        rows.append(row)
    if not rows:
        raise ValueError("识别到表头，但没有交易记录")
    return SourceData(headers, lookup, rows, sheet_name, provider, header_index)


def load_config(path: Path) -> Config:
    raw = json.loads(path.read_text(encoding="utf-8-sig"))
    categories = [
        Category(
            label=str(item["label"]).strip(),
            flow=str(item.get("flow") or "expense").strip().lower(),
            keywords=[str(keyword) for keyword in item.get("keywords", [])],
        )
        for item in raw["categories"]
    ]
    labels = [item.label for item in categories]
    if len(labels) != len(set(labels)) or any(not label for label in labels):
        raise ValueError("分类名称不能为空或重复")
    fallbacks = raw.get("fallbackByFlow") or {
        "expense": "其他支出", "income": "其他收入", "transfer": "其他转账"
    }
    for flow in ("expense", "income", "transfer"):
        if fallbacks.get(flow) not in labels:
            raise ValueError(f"缺少 {flow} 对应的兜底分类")
    return Config(categories, fallbacks, raw.get("merchantRules", []))


def load_overrides(path: Path, config: Config) -> dict[str, str]:
    """读取人工维护的“交易对方关键词 → 二级分类”映射。"""
    if not path.is_file():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(raw, dict):
        raise ValueError("merchant-overrides.json 必须是 JSON 对象")
    overrides: dict[str, str] = {}
    for merchant, category in raw.items():
        keyword = normalize_text(merchant)
        label = str(category).strip()
        if not keyword:
            raise ValueError("merchant-overrides.json 中的交易对方关键词不能为空")
        if label not in config.labels:
            raise ValueError(f"商户覆盖规则引用了不存在的分类：{label}")
        overrides[keyword] = label
    return overrides


def transaction_from_row(row: list[Any], lookup: dict[str, int]) -> dict[str, Any]:
    def get(key: str) -> Any:
        return row[lookup[key]] if lookup[key] >= 0 else None

    return {key: get(key) for key in (
        "time", "type", "counterparty", "account", "product", "direction",
        "amount", "payment", "status", "remark",
    )}


def transaction_text(transaction: dict[str, Any]) -> str:
    return normalize_text(" ".join(str(transaction.get(key) or "") for key in (
        "type", "counterparty", "account", "product", "direction", "status", "remark"
    )))


def transaction_flow(transaction: dict[str, Any]) -> str:
    text = transaction_text(transaction)
    if "退款" in text or "退货" in text:
        return "income"
    if any(marker in text for marker in ("收益发放", "利息收入", "结息", "存款利息")):
        return "income"
    transfer_markers = (
        "信用卡还款", "花呗还款", "花呗主动还款", "银行转账", "银行卡转账",
        "账户转账", "转入零钱通", "转出零钱通", "零钱提现", "微信提现",
        "转入余额宝", "余额宝转入", "转账收款到余额宝", "存款", "取款",
        "借入", "借出", "偿还借款", "收回借款", "收债", "垫付", "报销",
    )
    if any(marker in text for marker in transfer_markers):
        return "transfer"
    transaction_type = normalize_text(transaction.get("type"))
    if "转账" in transaction_type or "还款" in transaction_type:
        return "transfer"
    direction = normalize_text(transaction.get("direction"))
    if "收入" in direction or direction in {"收", "+"}:
        return "income"
    return "expense"


def classify(transaction: dict[str, Any], config: Config, overrides: dict[str, str]) -> Result:
    text = transaction_text(transaction)
    flow = transaction_flow(transaction)
    eligible = config.categories_for_flow(flow)
    allowed = config.labels_for_flow(flow)

    # 覆盖规则由用户明确维护，优先级最高。既支持完整交易对方，也支持“🏐”、
    # “KUMO KUMO”这样的稳定片段；多个片段同时命中时取最长者。
    counterparty = normalize_text(transaction.get("counterparty"))
    matched_overrides = [
        (keyword, category)
        for keyword, category in overrides.items()
        if category in allowed and (keyword == counterparty or keyword in counterparty)
    ]
    if matched_overrides:
        _, category = max(matched_overrides, key=lambda item: len(item[0]))
        return Result(category, 0.995, "商户覆盖规则")

    if ("退款" in text or "退货" in text) and "退款" in allowed:
        return Result("退款", 0.995, "退款规则")

    for rule in config.merchant_rules:
        keyword = normalize_text(rule.get("contains"))
        if keyword and keyword in text and rule.get("category") in allowed:
            confidence = float(rule.get("confidence", 0.95))
            if "群收款" in text:
                confidence = min(confidence, 0.58)
            return Result(str(rule["category"]), confidence, "商户规则")

    source_category = normalize_text(transaction.get("type"))
    hint = SOURCE_CATEGORY_HINTS.get(source_category)
    if hint and hint not in allowed:
        hint = None
    scored: list[tuple[float, Category, list[str]]] = []
    for category in eligible:
        keywords = [*category.keywords, *EXTRA_KEYWORDS.get(category.label, [])]
        matches = [keyword for keyword in keywords if normalize_text(keyword) and normalize_text(keyword) in text]
        score = sum(min(4.0, 1.5 + len(normalize_text(keyword)) * 0.25) for keyword in matches)
        if hint == category.label:
            score += 2.4
        scored.append((score, category, matches))
    scored.sort(key=lambda item: item[0], reverse=True)
    top_score, top_category, matches = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    second_matches = scored[1][2] if len(scored) > 1 else []

    if top_score == 0:
        return Result(config.fallback_by_flow[flow], 0.25, "无关键词")

    confidence = min(0.96, 0.80 + top_score * 0.035)
    top_specificity = max((len(normalize_text(keyword)) for keyword in matches), default=0)
    second_specificity = max((len(normalize_text(keyword)) for keyword in second_matches), default=0)
    if second_score > 0 and top_score - second_score < 1.2 and top_specificity < second_specificity + 2:
        confidence -= 0.18
    if "群收款" in text or "二维码收款" in text:
        confidence = min(confidence, 0.58)
    if hint and not matches:
        confidence = min(confidence, 0.80)
    return Result(top_category.label, max(0.30, confidence), "关键词规则")


def excel_column(index: int) -> str:
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def build_workbook(source: SourceData, results: list[Result], threshold: float) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "分类结果"
    sheet.append([*source.headers, *APPENDED_HEADERS])
    source_count = len(source.headers)
    id_indexes = {source.lookup[key] for key in ("transaction_id", "merchant_id") if source.lookup[key] >= 0}
    for source_row, result in zip(source.rows, results):
        sheet.append([*source_row, result.category, "是" if result.confidence < threshold else "否"])

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    body_font = Font(name="Microsoft YaHei", size=10)
    bottom_border = Border(bottom=Side(style="thin", color="E2E8F0"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            cell.border = bottom_border
        sheet.row_dimensions[row[0].row].height = 26
    for index in id_indexes:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, index + 1).number_format = "@"
    if source.lookup["time"] >= 0:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, source.lookup["time"] + 1).number_format = "yyyy-mm-dd hh:mm:ss"
    if source.lookup["amount"] >= 0:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, source.lookup["amount"] + 1)
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    default_widths = {
        "time": 22, "type": 16, "counterparty": 27, "account": 24, "product": 35,
        "direction": 11, "amount": 12, "payment": 16, "status": 16,
        "transaction_id": 34, "merchant_id": 34, "remark": 20,
    }
    for key, index in source.lookup.items():
        if index >= 0:
            sheet.column_dimensions[excel_column(index + 1)].width = default_widths.get(key, 16)
    sheet.column_dimensions[excel_column(source_count + 1)].width = 18
    sheet.column_dimensions[excel_column(source_count + 2)].width = 20
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, source_count + 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row, source_count + 2).alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    table = Table(displayName="LightClassificationResults", ref=f"A1:{excel_column(sheet.max_column)}{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)
    return workbook


def verify_output(path: Path, source: SourceData, labels: set[str]) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != ["分类结果"]:
            raise RuntimeError("输出应且仅应包含“分类结果”工作表")
        sheet = workbook["分类结果"]
        if sheet.max_row != len(source.rows) + 1 or sheet.max_column != len(source.headers) + 2:
            raise RuntimeError("输出行列数与输入不一致")
        if [sheet.cell(1, sheet.max_column - 1).value, sheet.cell(1, sheet.max_column).value] != APPENDED_HEADERS:
            raise RuntimeError("输出分类列标题不正确")
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, sheet.max_column - 1).value not in labels:
                raise RuntimeError(f"第 {row} 行分类不在清单中")
            if sheet.cell(row, sheet.max_column).value not in {"是", "否"}:
                raise RuntimeError(f"第 {row} 行人工确认标记错误")
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="微信/支付宝账单轻量分类器（无大模型、无交互）")
    parser.add_argument("input", type=Path, help="微信 .xlsx 或支付宝 .csv 文件")
    parser.add_argument("--output", type=Path, help="输出 .xlsx 路径")
    parser.add_argument("--categories", type=Path, default=DEFAULT_CATEGORIES, help="分类配置 JSON")
    parser.add_argument("--overrides", type=Path, default=DEFAULT_OVERRIDES, help="商户覆盖规则 JSON")
    parser.add_argument("--sheet", help="指定 Excel 工作表")
    parser.add_argument("--threshold", type=float, default=0.82, help="需要人工确认的置信度阈值，默认 0.82")
    args = parser.parse_args()
    if not 0 <= args.threshold <= 1:
        parser.error("--threshold 必须在 0 到 1 之间")
    return args


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    output_path = (args.output or input_path.parent / "outputs" / f"{input_path.stem}_已分类_轻量版.xlsx").resolve()
    config = load_config(args.categories.resolve())
    overrides = load_overrides(args.overrides.resolve(), config)
    source = clean_source(input_path, args.sheet)
    results = [classify(transaction_from_row(row, source.lookup), config, overrides) for row in source.rows]
    workbook = build_workbook(source, results, args.threshold)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    verify_output(output_path, source, set(config.labels))
    pending = sum(result.confidence < args.threshold for result in results)
    print(
        f"完成：{output_path}\n来源：{source.provider}；删除表头前内容 {source.removed_leading_rows} 行；"
        f"交易 {len(results)} 笔；需要人工确认 {pending} 笔。"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1) from exc
